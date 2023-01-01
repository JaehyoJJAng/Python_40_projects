import glob
import os
from openpyxl import load_workbook
import pandas as pd
from typing import List,Dict

class FindFiles:
    """ 파일 찾는 클래스 """
    def __init__(self,dir:str)-> None:
        self.dir = dir

    def get_xlsx_files(self)-> List[str]:
        """ 발주서.xlsx 파일만 추출해서 리턴하는 메서드 """
        return glob.glob(os.path.join(self.dir,'발주서_*.xlsx'))

class LoadXlsx:
    """ xlsx 데이터 읽는 클래스 """
    def __init__(self,files:list)-> None:
        self.files : list = files
    
    def load_data(self)-> List[Dict[str,str]]:
        area = []
        
        all_data = []

        # 발주처 파일 돌면서 데이터 가져오기
        for file in self.files:
            wb = load_workbook(file,data_only=True)
            ws = wb.active
            
            # 발주처 area에 추가
            area.append(ws['B1'].value)
            
            # A4 ~ B12 행까지의 데이터 가져오기
            for data in ws['A4':'B12']:
                data_dic = dict()

                if data[0].value != None and data[1].value != None:
                    all_data.append([data[0].value,data[1].value])
        return all_data

class SaveNewXlsx:
    """ 발주수량 합치는 클래스 """
    def __init__(self,all_data:list,dir:str)-> None:
        self.all_data : list = all_data
        self.dir = dir

    def combine(self)-> None:
        """ Pandas로 발주수량 합치기 """
        prod = []
        count = []
        for data in self.all_data:
            prod.append(data[0])
            count.append(data[1])

        df = pd.DataFrame({'물품':prod,'수량':count})
        df = df.groupby('물품').sum()

        # Excel Save
        df.to_excel(os.path.join(self.dir,'result.xlsx'))

def main()-> None:
    srcDir = './xlsx'

    # Get Files
    files : list = FindFiles(dir=srcDir).get_xlsx_files()
    
    # Get All data in Xlsx
    all_data : list = LoadXlsx(files=files).load_data()

    
    # Add All data
    SaveNewXlsx(all_data=all_data,dir=srcDir).combine()


if __name__ == '__main__':
    main()

