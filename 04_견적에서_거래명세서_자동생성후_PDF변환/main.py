""" pip install pywin32 """
import os
from openpyxl import load_workbook
from datetime import date

# ====== BASE DIRECOTRY ======
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# ======                ======
class Xlsx:
    def __init__(self,DirPath:str):
        self.DirPath = DirPath

    def result_price(self,count_list:list,unitCost_list:list)-> list:
        """ 가격 계산 메서드 """
        return [unitCost_list[idx] * count_list[idx] for idx in range(len(count_list))]

    def set_workbook(self,fileName:str)-> tuple:
        """ 워크북 세팅 """
        wb = load_workbook(os.path.join(self.DirPath,fileName))
        ws = wb.active
        return (wb,ws)

    def read_data_in_estimate(self)-> dict:
        """ 견적서 읽어오는 메소드 """
        # Save Dictionary
        save_dic = dict()

        # Work Sheet Instance
        wb,ws = self.set_workbook(fileName='견적서_샘플.xlsx')

        # 하나의 셀 읽기
        estimater = ws['A4'].value # 견적받는자
        a_total = ws['X25'].value # 소계
        surtax = ws['X26'].value # 부가세
        total_price = ws['X27'].value # 총합계 금액

        # 여러 개의 셀 반복하여 읽기
        prod_list = [] # 품목명
        count_list = [] # 수량
        unitCost_list = [] # 단가
        price_list = [] # 금액

        # ord : 문자를 아스키 번호로 변환하는 내장함수
        # ord("C")-64 == 3
        for data in ws['C13':'X24']:
            for cell in data:
                if cell.value != None:
                    if cell.column == ord('C')-64:
                        prod_list.append(cell.value)
                    elif cell.column == ord('R')-64:
                        count_list.append(cell.value)
                    elif cell.column == ord('T')-64:
                        unitCost_list.append(cell.value)
                    elif cell.column == ord('X')-64:
                        price_list.append(cell.value)

        # 가격 계산
        price_list : list = self.result_price(count_list=count_list,unitCost_list=unitCost_list)

        # save_dic 딕셔너리 변수에 데이터 집어넣기
        save_dic['estimater'] = estimater
        save_dic['prod_list'] = prod_list
        save_dic['count_list'] = count_list
        save_dic['unitCost_list'] = unitCost_list
        save_dic['price_list'] = price_list

        # Data Return
        return save_dic

    def write_data(self,estimate_datas:dict):
        """ 거래명세표에 값 기록하기 """
        wb,ws = self.set_workbook(fileName="거래명세표_샘플.xlsx")

        # 오늘 날짜 추출
        today = date.today() # 2022-12-23

        # C4 행에 연도 대입
        ws['C4'].value = today.year

        # E4 행에 월 대입
        ws['E4'].value = today.month

        # G4 행에 일 대입
        ws['G4'].value = today.day

        # C6 행에 견적 받는사람 대입
        ws['C6'].value = estimate_datas['estimater']

        for i in range(len(estimate_datas['prod_list'])):
            ws.cell(row=i+10,column=ord('B')-64).value = estimate_datas['prod_list'][i]
            ws.cell(row=i + 10, column=ord('G') - 64).value = estimate_datas['count_list'][i]
            ws.cell(row=i + 10, column=ord('I') - 64).value = estimate_datas['unitCost_list'][i]

        # Save
        self.save(wb=wb,estimater=estimate_datas['estimater'])

    def save(self,wb,estimater:str):
        """ 거래명세표 저장하기 """
        # 저장 경로
        savePath = os.path.abspath(f'result/{estimater}')
        if not os.path.exists(savePath):
            os.makedirs(savePath)

        # 파일 이름
        fileName = os.path.join(savePath,f'거래명세표_{estimater}.xlsx')

        # 거래명세표 저장
        wb.save(fileName)
        wb.close()

        # 출력
        print(f"저장 완료!\n{fileName}")

    def convert_pdf(self):
        """ 거래명세표 PDF로 변환
        pypiwin32 라이브러리가 설치가 안되는 관계로 Pass """
        pass

def main():
    DirPath = os.path.join(BASE_DIR,'xlsx')

    # Create XLSX Instance
    xlsx = Xlsx(DirPath=DirPath)
    estimate_infos: dict = xlsx.read_data_in_estimate()

    # 변경 데이터 새로 쓰기
    xlsx.write_data(estimate_datas=estimate_infos)

if __name__ == '__main__':
    main()



