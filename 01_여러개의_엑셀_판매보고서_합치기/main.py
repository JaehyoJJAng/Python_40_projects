import glob
import os 
from openpyxl import Workbook
from openpyxl import load_workbook

def load_files(dir:str)-> list:
    """ .xlsx 파일 모두 가져오는 함수 """
    files : list = [f for f in glob.glob(os.path.join(dir,'*.xlsx')) if "result" not in f]
    return files

def load_data_in_xlsx(file:str)-> dict:
    """ 엑셀 데이터 가져오는 함수 """
    wb = load_workbook(file,data_only=True)
    ws = wb.active
    
    data = {
            'area':ws['B1'].value,
            'dated':str(ws['B2'].value),
            'price':ws['B3'].value
            }
    return data
 
def create_result(datas:list,dir:str):
    """ 합친 엑셀 파일들의 데이터를 result.xlsx로 만들기 """

    # result.xslx 파일이 있다면 엑셀 파일 불러오고 아니면 새로운 워크북 객체 생성
    try:
        wb = load_workbook(os.path.join(dir,'result.xlsx'),data_only=True)
        ws = wb.active
    except:
        wb = Workbook()
        ws = wb.active
    
    # row 초기 값
    row = 1
    for data in datas:
        ws.cell(row=row,column=1).value = data['area']
        ws.cell(row=row,column=2).value = data['dated']
        ws.cell(row=row,column=3).value = data['price']
        
        # row 증감
        row += 1

    wb.save(os.path.join(dir,'result.xlsx'))
    wb.close()

def main():
    # Source Directory
    DIR = './xlsx'

    # Load Files
    files : list = load_files(dir=DIR)

    # Load Data
    result : list = [load_data_in_xlsx(file=file) for file in files]
    
    # Save new file
    create_result(datas=result,dir=DIR)

if __name__ == '__main__':
    main()
